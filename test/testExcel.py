from openpyxl import load_workbook
import pandas as pd

r = open('nganmach.py','r')
lines = r.readlines()
content = []
for line,value in enumerate(lines):
    content.append(value)
print(content)
df_new = pd.DataFrame({'E': content})
wb = load_workbook('test.xlsx')

ws = wb['Sheet1']

for index, row in df_new.iterrows():
    cell = 'E%d'  % (index + 1)
    ws[cell] = row[0]

wb.save('test.xlsx')